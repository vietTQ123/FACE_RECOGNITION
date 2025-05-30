import cv2
import numpy as np
import joblib
import pickle as pk
from skimage.feature import hog
from sklearn.decomposition import PCA
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ==== Duong dan chinh ====
dir_main = 'C:\\Users\\Admin\\Documents\\EMBEDDED_LAB\\FACE_RECOGNITION\\BTL_THCS\\code\\'

# ==== Tai mo hinh PCA va SVM ====
clf = joblib.load(f'{dir_main}c2_svm_classifier.pkl')
pca = pk.load(open(f"{dir_main}c1_PCA.pkl", 'rb'))
face_cascade = cv2.CascadeClassifier(f"{dir_main}haarcascade_frontalface_default.xml")

# ==== Tao file Excel neu chua ton tai ====
excel_path = f"{dir_main}diem_danh.xlsx"
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "DiemDanh"
    ws.append(["Ho va Ten", "Thoi gian diem danh"])
    wb.save(excel_path)

# ==== Map label => ten nguoi ====
label_map = {
    1: "Trinh_Quoc_Viet",
    3: "Le_Trung_Thanh",
    4: "Dao_Ha_Thai_Son",
}

# ==== Danh sach da diem danh trong phien ====
da_diem_danh = set()

# ==== Tham so he thong ====
confidence_threshold = 0.9

# ==== Khoi dong webcam ====
cam = cv2.VideoCapture(0)
if not cam.isOpened():
    print("Khong the mo camera.")
    exit()

print("[INFO] He thong san sang. Nhan 'q' de thoat.")

while True:
    ret, frame = cam.read()
    if not ret:
        print("Khong lay duoc khung hinh.")
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(60, 60))

    for (x, y, w, h) in faces:
        face_roi = gray[y:y+h, x:x+w]
        face_roi = cv2.equalizeHist(face_roi)
        resized = cv2.resize(face_roi, (128, 128))

        # Trich xuat dac trung HOG
        hog_features = hog(resized,
                           orientations=12,
                           pixels_per_cell=(4, 4),
                           cells_per_block=(3, 3),
                           block_norm='L2-Hys',
                           visualize=False)

        hog_features_pca = pca.transform([hog_features])
        probabilities = clf.predict_proba(hog_features_pca)[0]
        confidence = np.max(probabilities)
        label = clf.predict(hog_features_pca)[0]

        # Kiem tra dieu kien nhan dien
        if confidence >= confidence_threshold and label in label_map:
            label_text = label_map[label]
            color = (0, 255, 0)

            # Ghi diem danh neu chua ghi
            if label_text not in da_diem_danh:
                wb = load_workbook(excel_path)
                ws = wb["DiemDanh"]
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([label_text, now])
                wb.save(excel_path)
                da_diem_danh.add(label_text)
                print(f"[✔] Da diem danh: {label_text} luc {now}")
        else:
            label_text = "Unknown"
            color = (0, 0, 255)

        # Hien thi ket qua
        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, f"{label_text} ({confidence:.2f})", (x, y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.75, color, 2)
        print(f"[INFO] {label_text} | Confidence: {confidence:.2f}")

    cv2.imshow("Face Recognition", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cam.release()
cv2.destroyAllWindows()
